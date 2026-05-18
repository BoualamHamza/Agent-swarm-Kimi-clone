"""Recalculate every formula in an .xlsx file via headless LibreOffice and
report any Excel errors found. Single-file helper — drop into the sandbox at
/home/user/skills/xlsx/scripts/recalc.py and run:

    python recalc.py <path-to-xlsx> [timeout_seconds]

Returns a JSON report to stdout:
    {
      "status": "success" | "errors_found",
      "total_formulas": <int>,
      "total_errors":   <int>,
      "error_summary":  { "#REF!": {"count": N, "locations": [...]}, ... }
    }

Target environment: Linux (e2b sandbox). LibreOffice must be installed
(`apt-get install -y libreoffice`).
"""
from __future__ import annotations

import json
import os
import subprocess
import sys
from pathlib import Path

MACRO_DIR = os.path.expanduser("~/.config/libreoffice/4/user/basic/Standard")
MACRO_FILE = os.path.join(MACRO_DIR, "Module1.xba")

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

EXCEL_ERRORS = ("#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A")


def _soffice_env() -> dict:
    env = os.environ.copy()
    env.setdefault("SAL_USE_VCLPLUGIN", "svp")
    return env


def _setup_macro() -> bool:
    if os.path.exists(MACRO_FILE) and "RecalculateAndSave" in Path(MACRO_FILE).read_text():
        return True
    if not os.path.exists(MACRO_DIR):
        try:
            subprocess.run(
                ["soffice", "--headless", "--terminate_after_init"],
                capture_output=True, timeout=15, env=_soffice_env(),
            )
        except Exception:
            pass
        os.makedirs(MACRO_DIR, exist_ok=True)
    try:
        Path(MACRO_FILE).write_text(RECALCULATE_MACRO)
        return True
    except OSError:
        return False


def _scan_errors(filename: str) -> dict:
    from openpyxl import load_workbook

    wb_values = load_workbook(filename, data_only=True)
    locations: dict[str, list[str]] = {err: [] for err in EXCEL_ERRORS}
    total = 0
    for sheet_name in wb_values.sheetnames:
        ws = wb_values[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for err in EXCEL_ERRORS:
                        if err in cell.value:
                            locations[err].append(f"{sheet_name}!{cell.coordinate}")
                            total += 1
                            break
    wb_values.close()

    wb_formulas = load_workbook(filename, data_only=False)
    formula_count = 0
    for sheet_name in wb_formulas.sheetnames:
        ws = wb_formulas[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
    wb_formulas.close()

    summary = {
        err: {"count": len(locs), "locations": locs[:20]}
        for err, locs in locations.items()
        if locs
    }
    return {
        "status": "success" if total == 0 else "errors_found",
        "total_formulas": formula_count,
        "total_errors": total,
        "error_summary": summary,
    }


def recalc(filename: str, timeout: int = 60) -> dict:
    p = Path(filename)
    if not p.exists():
        return {"error": f"File {filename} does not exist"}
    if not _setup_macro():
        return {"error": "Failed to set up LibreOffice macro"}

    cmd = [
        "timeout", str(timeout),
        "soffice", "--headless", "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        str(p.absolute()),
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, env=_soffice_env())
    # rc 124 = timeout fired (file may still be calculated)
    if result.returncode not in (0, 124):
        return {
            "error": (result.stderr or "Unknown error during recalculation").strip()
        }

    try:
        return _scan_errors(filename)
    except Exception as e:
        return {"error": f"Error-scan failed: {e}"}


def main() -> int:
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Usage: recalc.py <xlsx_file> [timeout_seconds]"}))
        return 1
    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 60
    print(json.dumps(recalc(filename, timeout), indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
